# ╔══════════════════════════════════════════════════════════════╗
# ║                   phonebook_import.py                       ║
# ║  Импорт телефонного справочника из Excel                    ║
# ║                                                             ║
# ║  Маршруты:                                                  ║
# ║  GET  /phonebook/import/template  — скачать шаблон .xlsx    ║
# ║  POST /phonebook/import/upload    — загрузить файл (шаг 1)  ║
# ║  GET  /phonebook/import/resolve   — разрешить орг. (шаг 2)  ║
# ║  POST /phonebook/import/resolve   — сохранить маппинг       ║
# ║  POST /phonebook/import/confirm   — финальная загрузка (ш3) ║
# ╚══════════════════════════════════════════════════════════════╝

import os
import json
import uuid
from io import BytesIO
from datetime import datetime

from flask import (
    Blueprint, render_template, request,
    redirect, url_for, flash, send_file,
    session, jsonify
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db import get_db, UPLOADS_DIR
from auth_utils import login_required, admin_required
from activity_log import log_action

pb_import_bp = Blueprint('pb_import', __name__)

# ─── Константы ───────────────────────────────────────────────────────────────

# Заголовки колонок шаблона (порядок важен — используется при парсинге)
TEMPLATE_HEADERS = [
    'Организация',
    'ФИО',
    'Должность',
    'Кабинет',
    'Рабочий тел.',
    'Доб.',
    'Личный тел.',
    'Email',
    'Примечания',
]

# Папка для временных файлов импорта
IMPORT_TEMP_DIR = os.path.join(UPLOADS_DIR, 'import_temp')
os.makedirs(IMPORT_TEMP_DIR, exist_ok=True)


# ─── Вспомогательные функции ─────────────────────────────────────────────────

def _get_all_orgs():
    """Возвращает dict {name_lower: id} всех организаций."""
    conn = get_db()
    rows = conn.execute("SELECT id, name FROM phonebook_orgs ORDER BY name").fetchall()
    conn.close()
    return {r['name'].strip().lower(): r['id'] for r in rows}


def _parse_xlsx(filepath):
    """
    Разбирает загруженный Excel-файл.
    Возвращает (rows, errors):
      rows   — список dict с данными контактов
      errors — список строк с описанием ошибок
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return [], [f'Не удалось открыть файл: {e}']

    ws = wb.active
    errors = []
    rows = []

    # Ищем строку-заголовок (первая строка где первая ячейка == 'Организация')
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and str(row[0]).strip() == 'Организация':
            header_row_idx = i
            break

    if header_row_idx is None:
        return [], ['Не найдена строка заголовков. Убедитесь что используете шаблон SONAR.']

    # Проверяем соответствие заголовков
    header_cells = list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=True
    ))[0]
    found_headers = [str(c).strip() if c else '' for c in header_cells[:len(TEMPLATE_HEADERS)]]
    if found_headers != TEMPLATE_HEADERS:
        errors.append(
            f'Заголовки колонок не совпадают с шаблоном. '
            f'Ожидалось: {TEMPLATE_HEADERS}. Найдено: {found_headers}'
        )
        return [], errors

    # Читаем данные
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        cells = list(ws.iter_rows(
            min_row=row_idx, max_row=row_idx, values_only=True
        ))[0]

        # Пропускаем пустые строки
        if not any(cells[:len(TEMPLATE_HEADERS)]):
            continue

        full_name = str(cells[1]).strip() if cells[1] else ''
        if not full_name or full_name.lower() in ('фио', 'none', ''):
            errors.append(f'Строка {row_idx}: пустое поле «ФИО» — строка пропущена')
            continue

        rows.append({
            'org_name':       str(cells[0]).strip() if cells[0] else '',
            'full_name':      full_name,
            'position':       str(cells[2]).strip() if cells[2] else '',
            'room':           str(cells[3]).strip() if cells[3] else '',
            'phone_work':     str(cells[4]).strip() if cells[4] else '',
            'phone_ext':      str(cells[5]).strip() if cells[5] else '',
            'phone_personal': str(cells[6]).strip() if cells[6] else '',
            'email':          str(cells[7]).strip() if cells[7] else '',
            'notes':          str(cells[8]).strip() if cells[8] else '',
        })

    return rows, errors


def _save_import_session(data: dict) -> str:
    """Сохраняет данные импорта во временный JSON-файл, возвращает токен."""
    token = uuid.uuid4().hex
    path = os.path.join(IMPORT_TEMP_DIR, f'import_{token}.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    return token


def _load_import_session(token: str):
    """Загружает данные импорта по токену. Возвращает dict или None."""
    if not token:
        return None
    path = os.path.join(IMPORT_TEMP_DIR, f'import_{token}.json')
    if not os.path.exists(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def _delete_import_session(token: str):
    """Удаляет временный файл импорта."""
    path = os.path.join(IMPORT_TEMP_DIR, f'import_{token}.json')
    if os.path.exists(path):
        os.remove(path)


# ─── ШАГ 0: Скачать шаблон ───────────────────────────────────────────────────

@pb_import_bp.route('/phonebook/import/template')
@login_required
@admin_required
def import_template():
    """Генерирует и отдаёт шаблон Excel для заполнения."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Справочник'

    # Стили
    hfill = PatternFill('solid', fgColor='1B5E7B')
    hfont = Font(bold=True, color='FFFFFF', size=10)
    efill = PatternFill('solid', fgColor='FFF9C4')  # жёлтый — строка примера
    br = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Заголовок файла
    ws.merge_cells(f'A1:{get_column_letter(len(TEMPLATE_HEADERS))}1')
    ws['A1'].value = 'Шаблон импорта телефонного справочника SONAR'
    ws['A1'].font = Font(bold=True, size=12, color='1B5E7B')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f'A2:{get_column_letter(len(TEMPLATE_HEADERS))}2')
    ws['A2'].value = (
        'Не изменяйте заголовки колонок (строка 3). '
        'Строка 4 — пример, удалите перед загрузкой. '
        'Поле «ФИО» обязательно.'
    )
    ws['A2'].font = Font(italic=True, size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Заголовки колонок (строка 3)
    col_widths = [30, 28, 24, 10, 16, 8, 16, 24, 30]
    for ci, (h, w) in enumerate(zip(TEMPLATE_HEADERS, col_widths), 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = hfill
        c.font = hfont
        c.border = br
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 20

    # Пример строки (строка 4)
    example = [
        'Министерство инвестиций НО',
        'Иванов Иван Иванович',
        'Начальник отдела',
        '214',
        '+7 831 200-00-00',
        '101',
        '+7 912 000-00-00',
        'ivanov@example.ru',
        'Пример — удалите эту строку',
    ]
    for ci, val in enumerate(example, 1):
        c = ws.cell(row=4, column=ci, value=val)
        c.fill = efill
        c.border = br
        c.alignment = Alignment(vertical='center')
    ws.row_dimensions[4].height = 16

    ws.freeze_panes = 'A4'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'sonar_phonebook_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ─── ШАГ 1: Загрузка файла ───────────────────────────────────────────────────

@pb_import_bp.route('/phonebook/import/upload', methods=['POST'])
@login_required
@admin_required
def import_upload():
    """Принимает xlsx, парсит, определяет неизвестные организации."""
    f = request.files.get('import_file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('Загрузите файл в формате .xlsx', 'error')
        return redirect(url_for('phonebook.phonebook'))

    # Сохраняем временно
    tmp_name = f'upload_{uuid.uuid4().hex}.xlsx'
    tmp_path = os.path.join(IMPORT_TEMP_DIR, tmp_name)
    f.save(tmp_path)

    rows, errors = _parse_xlsx(tmp_path)
    os.remove(tmp_path)

    if errors and not rows:
        for e in errors:
            flash(e, 'error')
        return redirect(url_for('phonebook.phonebook'))

    if not rows:
        flash('Файл не содержит данных для импорта', 'error')
        return redirect(url_for('phonebook.phonebook'))

    # Определяем неизвестные организации
    known_orgs = _get_all_orgs()  # {name_lower: id}
    unknown_orgs = sorted(set(
        r['org_name'] for r in rows
        if r['org_name'] and r['org_name'].strip().lower() not in known_orgs
    ))

    # Сохраняем в сессию
    token = _save_import_session({
        'rows': rows,
        'parse_errors': errors,
        'unknown_orgs': unknown_orgs,
    })
    session['import_token'] = token

    if unknown_orgs:
        return redirect(url_for('pb_import.import_resolve'))
    else:
        return redirect(url_for('pb_import.import_confirm_get'))


# ─── ШАГ 2: Разрешение организаций ───────────────────────────────────────────

@pb_import_bp.route('/phonebook/import/resolve', methods=['GET'])
@login_required
@admin_required
def import_resolve():
    """Страница выбора: создать новую орг. или привязать к существующей."""
    token = session.get('import_token')
    data = _load_import_session(token)
    if not data:
        flash('Сессия импорта устарела. Загрузите файл заново.', 'error')
        return redirect(url_for('phonebook.phonebook'))

    conn = get_db()
    all_orgs = conn.execute(
        'SELECT id, name FROM phonebook_orgs ORDER BY name'
    ).fetchall()
    conn.close()

    return render_template(
        'phonebook_import_resolve.html',
        unknown_orgs=data['unknown_orgs'],
        all_orgs=all_orgs,
        rows_count=len(data['rows']),
        parse_errors=data.get('parse_errors', []),
    )


@pb_import_bp.route('/phonebook/import/resolve', methods=['POST'])
@login_required
@admin_required
def import_resolve_post():
    """
    Принимает маппинг организаций:
      - org_action_<slug>  = 'create' | 'map'
      - org_map_<slug>     = id существующей организации (если 'map')
      - org_new_name_<slug> = название новой орг. (если 'create', по умолчанию = оригинал)
    """
    token = session.get('import_token')
    data = _load_import_session(token)
    if not data:
        flash('Сессия импорта устарела. Загрузите файл заново.', 'error')
        return redirect(url_for('phonebook.phonebook'))

    mapping = {}  # {original_name: {'action': 'create'|'map', 'map_id': int|None, 'new_name': str}}
    for org_name in data['unknown_orgs']:
        slug = _org_slug(org_name)
        action = request.form.get(f'org_action_{slug}', 'create')
        map_id = request.form.get(f'org_map_{slug}', '')
        new_name = request.form.get(f'org_new_name_{slug}', org_name).strip() or org_name
        mapping[org_name] = {
            'action': action,
            'map_id': int(map_id) if map_id and map_id.isdigit() else None,
            'new_name': new_name,
        }

    data['org_mapping'] = mapping
    _save_import_session_by_token(token, data)
    return redirect(url_for('pb_import.import_confirm_get'))


# ─── ШАГ 3: Предпросмотр и финальная загрузка ────────────────────────────────

@pb_import_bp.route('/phonebook/import/confirm', methods=['GET'])
@login_required
@admin_required
def import_confirm_get():
    """Показывает итоговый предпросмотр перед записью в БД."""
    token = session.get('import_token')
    data = _load_import_session(token)
    if not data:
        flash('Сессия импорта устарела. Загрузите файл заново.', 'error')
        return redirect(url_for('phonebook.phonebook'))

    return render_template(
        'phonebook_import_resolve.html',
        unknown_orgs=[],
        all_orgs=[],
        rows_count=len(data['rows']),
        parse_errors=data.get('parse_errors', []),
        preview_mode=True,
        rows_preview=data['rows'][:10],  # показываем первые 10 строк
    )


@pb_import_bp.route('/phonebook/import/confirm', methods=['POST'])
@login_required
@admin_required
def import_confirm_post():
    """Финальная запись данных в БД."""
    token = session.get('import_token')
    data = _load_import_session(token)
    if not data:
        flash('Сессия импорта устарела. Загрузите файл заново.', 'error')
        return redirect(url_for('phonebook.phonebook'))

    rows = data['rows']
    org_mapping = data.get('org_mapping', {})

    conn = get_db()
    known_orgs = {r['name'].strip().lower(): r['id']
                  for r in conn.execute('SELECT id, name FROM phonebook_orgs').fetchall()}

    created = 0
    skipped = 0
    orgs_created = 0

    try:
        # 1. Создаём новые организации из маппинга
        for orig_name, info in org_mapping.items():
            if info['action'] == 'create':
                new_name = info['new_name']
                key = new_name.strip().lower()
                if key not in known_orgs:
                    cur = conn.execute(
                        'INSERT INTO phonebook_orgs (name) VALUES (?)', (new_name,)
                    )
                    known_orgs[key] = cur.lastrowid
                    orgs_created += 1
            elif info['action'] == 'map' and info['map_id']:
                # Привязываем оригинальное имя к существующей org
                known_orgs[orig_name.strip().lower()] = info['map_id']

        # 2. Вставляем контакты
        for r in rows:
            org_name = r['org_name'].strip()
            org_key = org_name.lower()

            # Если орг. есть в маппинге и там 'map' — используем map_id
            if org_name in org_mapping and org_mapping[org_name]['action'] == 'map':
                org_id = org_mapping[org_name]['map_id']
            else:
                org_id = known_orgs.get(org_key)

            # Проверка дубликатов по ФИО + org_id
            exists = conn.execute(
                'SELECT id FROM phonebook WHERE full_name=? AND org_id IS ?',
                (r['full_name'], org_id)
            ).fetchone()
            if exists:
                skipped += 1
                continue

            conn.execute("""
                INSERT INTO phonebook
                    (org_id, full_name, position, room,
                     phone_work, phone_ext, phone_personal, email, notes)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (
                org_id,
                r['full_name'],
                r['position'],
                r['room'],
                r['phone_work'],
                r['phone_ext'],
                r['phone_personal'],
                r['email'],
                r['notes'],
            ))
            created += 1

        log_action(conn, session['user_id'], 'create', None,
                   f'Импорт справочника: добавлено {created} контактов, '
                   f'пропущено {skipped} дублей, создано {orgs_created} орг.')
        conn.commit()

    except Exception as e:
        conn.rollback()
        conn.close()
        _delete_import_session(token)
        session.pop('import_token', None)
        flash(f'Ошибка при импорте: {e}', 'error')
        return redirect(url_for('phonebook.phonebook'))

    conn.close()
    _delete_import_session(token)
    session.pop('import_token', None)

    flash(
        f'Импорт завершён: добавлено {created} контактов'
        + (f', пропущено {skipped} дублей' if skipped else '')
        + (f', создано {orgs_created} новых организаций' if orgs_created else ''),
        'success'
    )
    return redirect(url_for('phonebook.phonebook'))


# ─── AJAX: поиск организаций для datalist ────────────────────────────────────

@pb_import_bp.route('/phonebook/import/orgs_search')
@login_required
def orgs_search():
    """AJAX: возвращает список организаций для живого поиска."""
    q = request.args.get('q', '').strip()
    conn = get_db()
    if q:
        rows = conn.execute(
            "SELECT id, name FROM phonebook_orgs WHERE name LIKE ? ORDER BY name LIMIT 30",
            (f'%{q}%',)
        ).fetchall()
    else:
        rows = conn.execute(
            'SELECT id, name FROM phonebook_orgs ORDER BY name LIMIT 50'
        ).fetchall()
    conn.close()
    return jsonify([{'id': r['id'], 'name': r['name']} for r in rows])


# ─── Утилиты ─────────────────────────────────────────────────────────────────

def _org_slug(name: str) -> str:
    """Создаёт безопасный slug из названия организации для имён полей формы."""
    import re
    return re.sub(r'[^a-zA-Z0-9а-яА-ЯёЁ]', '_', name)[:60]


def _save_import_session_by_token(token: str, data: dict):
    """Обновляет существующий временный файл импорта."""
    path = os.path.join(IMPORT_TEMP_DIR, f'import_{token}.json')
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
