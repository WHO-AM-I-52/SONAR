# ╔══════════════════════════════════════════════════════════════╗
# ║               tools/investmap_export.py                     ║
# ║  Конвертер выгрузки ГИС НСИ (инвестплощадки) → текст + dict ║
# ║  Формат 1: 1 площадка (атрибут|значение по строкам)         ║
# ║  Формат 2: N площадок (строки=площадки, столбцы=атрибуты)   ║
# ║  Формат 3: 1 площадка ГИС НСИ (шапка на стр.2, 3 столбца)  ║
# ╚══════════════════════════════════════════════════════════════╝

import openpyxl
import io
import re
import html

# Поля, которые ПОЛНОСТЬЮ исключаются из data и не передаются анализатору:
# (даты, типы изменений, пользователь — технический мусор СИСТЕМЫ)
SERVICE_FIELDS = {
    'код во внешнем источнике',
    'тип последнего изменения',
    'дата последнего изменения',
    'дата создания',
    'пользователь/система, производивший изменение',
    'пользователь/система, производивший изменения',
    # global_id — НЕ фильтруем: он нужен анализатору для идентификации площадки в SMS
}

# ─────────────────────────────────────────────────────────────────────────────
# Нормализация заголовков портальной выгрузки (формат 2) → ключи анализатора.
# Портал даёт длинные/другие имена, анализатор ждёт короткие.
# ─────────────────────────────────────────────────────────────────────────────
PORTAL_HEADER_MAP = {
    'Фотографии объекта':                                                          'Фотографии',
    'Форма собственности объекта':                                                 'Форма собственности',
    'Стоимость объекта, руб. (покупки или месячной аренды)':                      'Стоимость объекта',
    'Стоимость, руб./год за га':                                                   'Стоимость руб./год за га',
    'Стоимость, руб./год за кв.м.':                                               'Стоимость руб./год за кв.м',
    'min и max сроки аренды (если применимо), лет':                               'Сроки аренды',
    'Класс опасности объекта':                                                     'Класс опасности',
    'Характеристики расположенных объектов капитального строительства':            'Характеристики ОКС',
    'Свободная площадь ЗУ, га':                                                   'Свободная площадь ЗУ',
    'Межевание ЗУ':                                                               'Межевание',
    'Свободная площадь здания, сооружения, помещения, кв. м':                     'Свободная площадь здания',
    'Кадастровый номер здания, сооружения, помещения':                            'Кадастровый номер здания',
    'Технические характеристики здания, сооружения, помещения':                   'Технические характеристики здания',
    'Наименование собственника / администратора объекта':                         'Наименование собственника',
    'Телефон контактного лица, e-mail':                                           'Телефон',
    'Водоснабжение Наличие (Да/Нет)':                                             'Водоснабжение Наличие',
    'Водоотведение Наличие (Да/Нет)':                                             'Водоотведение Наличие',
    'Газоснабжение Наличие (Да/Нет)':                                             'Газоснабжение Наличие',
    'Электроснабжение Наличие (Да/Нет)':                                          'Электроснабжение Наличие',
    'Теплоснабжение Наличие (Да/Нет)':                                            'Теплоснабжение Наличие',
    'Вывоз ТКО Наличие (Да/Нет)':                                                 'Вывоз ТКО Наличие',
    'Наличие подъездных путей (Да/Нет)':                                          'Наличие подъездных путей',
    'Наличие ж/д (Да/Нет)':                                                       'Наличие ж/д',
    'Наличие парковки грузового транспорта':                                       'Наличие парковки',
    'Перечень документов, необходимых для подачи заявки':                         'Перечень документов',
    'Перечень видов экономической деятельности, возможных к реализации на площадке': 'Перечень видов экономической',
    'Градостроительные характеристики и ограничения':                             'Градостроительные характеристики',
    'Статус площадки':                                                             'Статус',
    'Объекты водоснабжения Максимально допустимая мощность, куб. м/ч':            'Водоснабжение Максимально допустимая мощность',
    'Объекты водоснабжения Свободная мощность, куб.м/ч':                          'Водоснабжение Свободная мощность',
    'Объекты водоотведения Максимально допустимая мощность, куб. м/ч':            'Водоотведение Максимально допустимая мощность',
    'Объекты водоотведения Свободная мощность, куб.м/ч':                          'Водоотведение Свободная мощность',
    'Объекты газоснабжения Свободная мощность, куб. м/ч':                         'Газоснабжение Свободная мощность',
    'Объекты газоснабжения Величина максимального расхода газа (мощности) газоиспользующего оборудования, куб. м./ч': 'Газоснабжение Максимально допустимая мощность',
    'Объекты электроснабжения Максимальная мощность, МВт/ч':                      'Электроснабжение Максимальная мощность',
    'Объекты электроснабжения Свободная мощность, МВт/ч':                         'Электроснабжение Свободная мощность',
}

# Столбцы координат — особый случай: 4 портальных столбца → 3 ключа анализатора.
# Значения широты и долготы объединяются через запятую → 'Координаты (точка)'
PORTAL_COORD_LAT  = 'Широта объекта в координатах WGS-84'
PORTAL_COORD_LON  = 'Долгота объекта в координатах WGS-84'
PORTAL_COORD_LINE = 'Набор координат линии объекта в координатах WGS-84'
PORTAL_COORD_POLY = 'Набор координат полигона объекта в координатах WGS-84'

PORTAL_COORD_SKIP = {PORTAL_COORD_LAT, PORTAL_COORD_LON, PORTAL_COORD_LINE, PORTAL_COORD_POLY}


def _normalize_portal_header(h):
    """Приводит заголовок портального xlsx к ключу анализатора.
    Координатные столбцы обрабатываются отдельно в parse_format2,
    поэтому здесь возвращаем None для них — признак «пропустить в общем цикле».
    """
    if h in PORTAL_COORD_SKIP:
        return None   # обработаем вручную
    return PORTAL_HEADER_MAP.get(h, h)


def _is_portal_format2(headers):
    """Проверяем, что f2-файл является портальной выгрузкой по характерным заголовкам."""
    header_set = set(headers)
    portal_markers = {PORTAL_COORD_LAT, PORTAL_COORD_LON, 'Статус площадки', 'Формат площадки'}
    return len(portal_markers & header_set) >= 2


def _clean(val):
    """Нормализация значения ячейки: HTML-entities, теги, пробелы."""
    if val is None:
        return ''
    s = str(val).strip()
    s = re.sub(r'<[^>]+>', '', s)
    s = html.unescape(s)
    s = s.replace('\u00a0', ' ').strip()
    return s


def _is_archive(attr_name):
    """Поле с (архив) в названии — полностью игнорировать."""
    return '(архив)' in attr_name.lower()


def _is_service(attr_name):
    """Служебное поле системы — исключить из data."""
    return attr_name.lower().strip() in SERVICE_FIELDS


def _to_empty(val):
    """Пустые/недопустимые значения → ПУСТО."""
    if not val:
        return 'ПУСТО'
    lower = val.lower().strip()
    if lower in ('-', '—', 'нет данных', 'н/д', 'не указано', 'не заполнено'):
        return 'ПУСТО'
    return val


def _detect_format(ws):
    max_col = ws.max_column
    max_row = ws.max_row

    if max_col <= 4 and max_row >= 3:
        row2 = [_clean(c.value).lower() for c in next(ws.iter_rows(min_row=2, max_row=2))]
        row2_text = ' '.join(row2)
        if 'атрибут' in row2_text or 'значени' in row2_text:
            return 'f3'

    if max_col <= 2 and max_row >= 3:
        return 'f1'

    return 'f2'


def _build_output(attr, val):
    """Возвращает (текстовая строка, val_normalized) или None если поле пропустить."""
    if _is_archive(attr) or _is_service(attr):
        return None
    val_norm = _to_empty(_clean(val) if val else '')
    return f"{attr} → {val_norm}", val_norm


def parse_format1(ws):
    lines = []
    data = {}
    for row in ws.iter_rows(min_row=1):
        cells = [_clean(c.value) for c in row]
        attr = cells[0] if len(cells) > 0 else ''
        val  = cells[1] if len(cells) > 1 else ''
        if not attr:
            continue
        result = _build_output(attr, val)
        if result is None:
            continue
        line, val_norm = result
        lines.append(line)
        data[attr] = val_norm
    return lines, data


def parse_format3(ws):
    """
    Формат 3: 1 площадка ГИС НСИ.
    Строка 1 — заголовок каталога (пропускаем).
    Строка 2 — шапка колонок (пропускаем).
    Строки 3+: col A = атрибут, col B = описание, col C = значение.
    """
    lines = []
    data = {}
    for row in ws.iter_rows(min_row=3):
        cells = [_clean(c.value) for c in row]
        attr = cells[0] if len(cells) > 0 else ''
        val  = cells[2] if len(cells) > 2 else ''
        if not attr:
            continue
        result = _build_output(attr, val)
        if result is None:
            continue
        line, val_norm = result
        lines.append(line)
        data[attr] = val_norm
    return lines, data


def parse_format2(ws):
    """
    Формат 2: N площадок.
    Строка 1 = заголовки атрибутов. Строки 2+ = данные.

    Если файл является портальной выгрузкой — заголовки нормализуются
    через PORTAL_HEADER_MAP, а 4 координатных столбца объединяются в 3 ключа.
    """
    raw_headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        raw_headers.append(_clean(cell.value))

    is_portal = _is_portal_format2(raw_headers)

    # Строим индексы координатных столбцов (только для портала)
    idx_lat  = raw_headers.index(PORTAL_COORD_LAT)  if is_portal and PORTAL_COORD_LAT  in raw_headers else None
    idx_lon  = raw_headers.index(PORTAL_COORD_LON)  if is_portal and PORTAL_COORD_LON  in raw_headers else None
    idx_line = raw_headers.index(PORTAL_COORD_LINE) if is_portal and PORTAL_COORD_LINE in raw_headers else None
    idx_poly = raw_headers.index(PORTAL_COORD_POLY) if is_portal and PORTAL_COORD_POLY in raw_headers else None

    # Нормализуем заголовки
    if is_portal:
        headers = [_normalize_portal_header(h) for h in raw_headers]
    else:
        headers = raw_headers

    blocks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cells = [_clean(c.value) for c in row]
        if not any(cells):
            continue
        lines = [f"=== ПЛОЩАДКА {row_idx - 1} ==="]
        data = {}

        # Координаты (портал): объединяем широту+долготу → 'Координаты (точка)'
        if is_portal:
            lat      = cells[idx_lat]  if idx_lat  is not None and idx_lat  < len(cells) else ''
            lon      = cells[idx_lon]  if idx_lon  is not None and idx_lon  < len(cells) else ''
            line_val = cells[idx_line] if idx_line is not None and idx_line < len(cells) else ''
            poly_val = cells[idx_poly] if idx_poly is not None and idx_poly < len(cells) else ''

            if lat and lon:
                coord_point = f"{lat}, {lon}"
                data['Координаты (точка)'] = coord_point
                lines.append(f"Координаты (точка) → {coord_point}")
            if line_val:
                data['Координаты (линия)'] = _to_empty(line_val)
                lines.append(f"Координаты (линия) → {_to_empty(line_val)}")
            if poly_val:
                data['Координаты (полигон)'] = _to_empty(poly_val)
                lines.append(f"Координаты (полигон) → {_to_empty(poly_val)}")

        for h, v in zip(headers, cells):
            if not h:   # None — координатные столбцы портала, уже обработаны выше
                continue
            result = _build_output(h, v)
            if result is None:
                continue
            line, val_norm = result
            lines.append(line)
            data[h] = val_norm

        blocks.append((lines, data))
    return blocks


def convert_excel_to_text(file_bytes):
    """
    Основная функция.
    Принимает bytes файла .xlsx.
    Возвращает dict:
      {
        'format': 1, 2 или 3,
        'count': количество площадок,
        'text': итоговый текст для вставки в чат,
        'data': dict {атрибут: значение} для f1/f3, list[dict] для f2,
        'error': None или строка ошибки
      }
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        fmt = _detect_format(ws)

        if fmt == 'f3':
            lines, data = parse_format3(ws)
            return {'format': 3, 'count': 1, 'text': '\n'.join(lines), 'data': data, 'error': None}

        elif fmt == 'f1':
            lines, data = parse_format1(ws)
            return {'format': 1, 'count': 1, 'text': '\n'.join(lines), 'data': data, 'error': None}

        else:
            blocks = parse_format2(ws)
            text = '\n\n'.join(['\n'.join(b[0]) for b in blocks])
            data_list = [b[1] for b in blocks]
            return {'format': 2, 'count': len(blocks), 'text': text, 'data': data_list, 'error': None}

    except Exception as e:
        return {'format': None, 'count': 0, 'text': '', 'data': {}, 'error': str(e)}
