# ╔══════════════════════════════════════════════════════════════╗
# ║                   investmap_routes.py                        ║
# ║  v1.0: загрузка .xlsx инвестплощадки → текстовый блок        ║
# ╚══════════════════════════════════════════════════════════════╝

import io
from flask import Blueprint, render_template, request, jsonify, session
from auth_utils import login_required, admin_required

investmap_bp = Blueprint('investmap', __name__)


def _cell(val) -> str:
    """Возвращает строковое представление ячейки или пустую строку."""
    if val is None:
        return ''
    return str(val).strip()


def _parse_single(ws) -> str:
    """
    Формат 1: два столбца — «атрибут | значение».
    Возвращает текстовый блок.
    """
    lines = []
    for row in ws.iter_rows(values_only=True):
        key = _cell(row[0] if len(row) > 0 else None)
        val = _cell(row[1] if len(row) > 1 else None)
        if not key and not val:
            continue
        if key and val:
            lines.append(f"{key}: {val}")
        elif key:
            lines.append(key)
        elif val:
            lines.append(val)
    return '\n'.join(lines)


def _parse_multi(ws) -> str:
    """
    Формат 2: первая строка — заголовки, остальные — площадки.
    Каждая площадка выводится блоком «Заголовок: Значение».
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return ''

    headers = [_cell(h) for h in rows[0]]
    blocks = []

    for i, row in enumerate(rows[1:], start=1):
        parts = []
        for j, h in enumerate(headers):
            val = _cell(row[j] if j < len(row) else None)
            if not h and not val:
                continue
            if h and val:
                parts.append(f"{h}: {val}")
            elif val:
                parts.append(val)
        if parts:
            blocks.append(f"--- Площадка {i} ---\n" + '\n'.join(parts))

    return '\n\n'.join(blocks)


def _detect_format(ws) -> str:
    """
    Эвристика: если в первой строке обе ячейки выглядят как «атрибут | значение»
    (не числа, вторая колонка ≠ первая) и всего 2 значимых столбца → Формат 1.
    Иначе → Формат 2 (таблица).
    """
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    if not rows:
        return 'single'

    # Считаем максимальное число непустых столбцов в первых 3 строках
    max_cols = 0
    for row in rows:
        cnt = sum(1 for c in row if c is not None and str(c).strip())
        max_cols = max(max_cols, cnt)

    # Если столбцов больше 2 → явно таблица
    if max_cols > 2:
        return 'multi'

    return 'single'


@investmap_bp.route('/investmap', methods=['GET'])
@login_required
@admin_required
def investmap():
    return render_template('investmap.html')


@investmap_bp.route('/investmap/parse', methods=['POST'])
@login_required
@admin_required
def investmap_parse():
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'Файл не выбран'}), 400

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Поддерживаются только файлы .xlsx и .xls'}), 400

    try:
        import openpyxl
    except ImportError:
        return jsonify({'error': 'Библиотека openpyxl не установлена. Запустите install.bat.'}), 500

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'Не удалось открыть файл: {e}'}), 400

    fmt = request.form.get('format') or _detect_format(ws)

    if fmt == 'multi':
        result = _parse_multi(ws)
    else:
        result = _parse_single(ws)

    if not result.strip():
        return jsonify({'error': 'Файл пустой или не удалось распознать данные'}), 400

    return jsonify({'text': result, 'format': fmt})
