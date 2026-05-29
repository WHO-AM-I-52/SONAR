# ╔══════════════════════════════════════════════════════════════╗
# ║                       export_routes.py                       ║
# ║  v2.6: +export/full (полная выгрузка), +import/full (импорт) ║
# ╚══════════════════════════════════════════════════════════════╝

from flask import Blueprint, request, send_file, jsonify, session
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

from db import get_db, REPORTS_DIR
from auth_utils import login_required
from activity_log import log_action

report_bp = Blueprint('report', __name__)


# ─── ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ───────────────────────────────────────────────

def _short_fio(full_name: str) -> str:
    if not full_name:
        return '—'
    parts = full_name.strip().split()
    if len(parts) >= 3:
        return f"{parts[0]} {parts[1][0].upper()}.{parts[2][0].upper()}."
    if len(parts) == 2:
        return f"{parts[0]} {parts[1][0].upper()}."
    return full_name


def _std_border(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _hex_to_argb(hex_color: str) -> str:
    """'#RRGGBB', 'RRGGBB', '#AARRGGBB' → 'FFRRGGBB' (ARGB без #)."""
    c = hex_color.lstrip('#').upper() if hex_color else ''
    if len(c) == 6:
        return 'FF' + c
    if len(c) == 8:
        return c
    return ''   # пустая — значит не красить


def _fmt_date(iso) -> str:
    """Преобразует 'YYYY-MM-DD' или datetime → 'ДД.ММ.ГГГГ'."""
    if not iso:
        return '—'
    if isinstance(iso, (date, datetime)):
        return iso.strftime('%d.%m.%Y')
    s = str(iso).strip()[:10]
    try:
        return datetime.strptime(s, '%Y-%m-%d').strftime('%d.%m.%Y')
    except Exception:
        return s or '—'


def _mln_to_mld(val) -> str:
    if val is None or val == '':
        return '—'
    try:
        mld = float(str(val).replace(',', '.')) / 1000
        return f"{mld:.3f}".rstrip('0').rstrip('.')
    except (ValueError, TypeError):
        return str(val)


def _contact_cell(person: str, phone: str, email: str) -> str:
    parts = []
    if person and person.strip():
        parts.append(f"Ф.И.О. {person.strip()}")
    if phone and phone.strip():
        parts.append(f"Тел: {phone.strip()}")
    if email and email.strip():
        parts.append(email.strip())
    return '\n'.join(parts) if parts else '—'


# ─── СТАНДАРТНАЯ ВЫГРУЗКА ───────────────────────────────────────────────────

@report_bp.route('/report')
@login_required
def report():
    df = request.args.get('date_from', '')
    dt = request.args.get('date_to', '')
    sf = request.args.get('status', '')

    conn = get_db()
    q = ("SELECT r.*,u.full_name as employee_name,ass.full_name as assigned_name "
         "FROM requests r "
         "LEFT JOIN users u   ON r.created_by=u.id "
         "LEFT JOIN users ass ON r.assigned_to=ass.id "
         "WHERE 1=1")
    p = []
    if df:
        q += ' AND r.request_date>=?'; p.append(df)
    if dt:
        q += ' AND r.request_date<=?'; p.append(dt)
    if sf:
        q += ' AND r.status=?'; p.append(sf)

    rows = conn.execute(q + ' ORDER BY r.request_date', p).fetchall()

    sm = {
        'draft':    'Черновик',
        'review':   'На проверке',
        'accepted': 'Принято в работу',
        'answered': 'Ответ направлен',
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Отчёт'

    hfill = PatternFill("solid", fgColor="1B5E7B")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    alt   = PatternFill("solid", fgColor="EAF4FB")
    br    = _std_border()

    ws.merge_cells('A1:Q1')
    per = f" за период {_fmt_date(df)}–{_fmt_date(dt)}" if (df or dt) else ""
    ws['A1'].value     = f"Обращения на подбор земельных участков{per}"
    ws['A1'].font      = Font(bold=True, size=13, color="1B5E7B")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:Q2')
    ws['A2'].value = (
        f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}  "
        f"Всего: {len(rows)}"
    )
    ws['A2'].font      = Font(italic=True, size=9, color="888888")
    ws['A2'].alignment = Alignment(horizontal='center')

    hdrs = [
        '№ обращения', 'Дата', 'Статус', 'Источник', 'Заявитель', 'Название проекта',
        'Контактное лицо', 'Телефон', 'E-mail', 'Инвестиции (млн)',
        'Рабочих мест', 'Площадь (га)', 'Застройка (м²)', 'Право пользования',
        'Районы', 'Ответственный', 'Дата ответа',
    ]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = hfill; c.font = hfont; c.border = br
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 38

    for ri, r in enumerate(rows, 4):
        fill = alt if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [
            r['request_number'] or '—',
            _fmt_date(r['request_date']),
            sm.get(r['status'], r['status']),
            r['source_type'] or '—',
            r['applicant_short_name'] or r['applicant_full_name'] or '—',
            r['project_name'] or '—',
            r['contact_person'] or '—',
            r['contact_phone'] or '—',
            r['contact_email'] or '—',
            r['investment_total'],
            r['jobs_total'],
            r['site_area_ha'],
            r['site_build_area_m2'],
            r['site_right'] or '—',
            r['preferred_districts'] or '—',
            r['assigned_name'] or r['employee_name'] or '—',
            _fmt_date(r['answer_date']),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill; c.border = br
            c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[ri].height = 16

    for ci, w in enumerate(
        [16, 12, 20, 16, 28, 30, 20, 15, 24, 12, 10, 10, 12, 16, 24, 20, 12], 1
    ):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    fn = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fp = os.path.join(REPORTS_DIR, fn)
    wb.save(fp)

    # ── Журнал ──
    log_parts = []
    if df or dt:
        log_parts.append(f"период: {_fmt_date(df)} – {_fmt_date(dt)}")
    if sf:
        log_parts.append(f"статус: {sm.get(sf, sf)}")
    log_parts.append(f"всего {len(rows)} обращ.")
    try:
        log_action(conn, session['user_id'], 'export_report',
                   detail='; '.join(log_parts))
        conn.commit()
    except Exception:
        pass
    conn.close()

    return send_file(fp, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── ЕЖЕНЕДЕЛЬНАЯ ВЫГРУЗКА ДЛЯ МИНЭК ─────────────────────────────────────────

@report_bp.route('/report/minek')
@login_required
def report_minek():
    today      = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end   = week_start + timedelta(days=6)

    df = request.args.get('date_from', week_start.isoformat())
    dt = request.args.get('date_to',   week_end.isoformat())
    sf = request.args.get('status', '')

    conn = get_db()

    q = """
        SELECT
            r.*,
            u.full_name   AS employee_name,
            ass.full_name AS assigned_name,
            st.name       AS subject_type_name,
            rt.name       AS result_type_name,
            rt.color_hex  AS result_color
        FROM requests r
        LEFT JOIN users         u   ON r.created_by      = u.id
        LEFT JOIN users         ass ON r.assigned_to     = ass.id
        LEFT JOIN subject_types st  ON r.subject_type_id = st.id
        LEFT JOIN result_types  rt  ON r.result_type_id  = rt.id
        WHERE r.request_date >= ? AND r.request_date <= ?
    """
    p = [df, dt]
    if sf:
        q += ' AND r.status = ?'; p.append(sf)
    q += ' ORDER BY r.request_date, r.id'

    rows = conn.execute(q, p).fetchall()
    result_types = conn.execute(
        'SELECT id, name, color_hex FROM result_types ORDER BY id'
    ).fetchall()

    HEADER_COLOR = '1B5E7B'
    hfill = PatternFill('solid', fgColor=HEADER_COLOR)
    hfont = Font(bold=True, color='FFFFFF', size=10)
    alt   = PatternFill('solid', fgColor='EAF4FB')
    br    = _std_border()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Заявки'

    NCOLS = 12

    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    ws['A1'].value = (
        f"Еженедельный доклад МинЭК: обращения за период "
        f"{_fmt_date(df)} – {_fmt_date(dt)}"
    )
    ws['A1'].font      = Font(bold=True, size=13, color=HEADER_COLOR)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
    ws['A2'].value = (
        f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}   "
        f"Обращений в выборке: {len(rows)}"
    )
    ws['A2'].font      = Font(italic=True, size=9, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')

    HEADERS = [
        '',
        'Дата обращения',
        'Наименование компании',
        'Наименование проекта',
        'Объем инвестиций,\nмлрд рублей',
        'Рабочие места',
        'Предмет обращения',
        'Дата направления презентации',
        'Дата получения обратной связи',
        'Итоги работы по обращению',
        'Менеджер',
        'Телефон, контактное лицо',
    ]

    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = hfill; c.font = hfont; c.border = br
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 40

    for ri, r in enumerate(rows, 4):
        # Цвет строки: result_color хранится как 'RRGGBB' (без #)
        argb = _hex_to_argb(r['result_color']) if r['result_color'] else ''
        if argb:
            rfill = PatternFill('solid', fgColor=argb)
        else:
            rfill = alt if ri % 2 == 0 else PatternFill('solid', fgColor='FFFFFFFF')

        result_val = r['additional_info'] or r['result_type_name'] or '—'

        vals = [
            ri - 3,
            _fmt_date(r['request_date']),
            r['applicant_short_name'] or r['applicant_full_name'] or '—',
            r['project_name'] or '—',
            _mln_to_mld(r['investment_total']),
            r['jobs_total'] or '—',
            r['subject_type_name'] or '—',
            _fmt_date(r['answer_date']),
            _fmt_date(r['feedback_date']),
            result_val,
            _short_fio(r['assigned_name'] or r['employee_name']),
            _contact_cell(
                r['contact_person'],
                r['contact_phone'],
                r['contact_email'],
            ),
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill   = rfill
            c.border = br
            c.alignment = Alignment(
                vertical='center',
                wrap_text=True,
                horizontal='center' if ci == 1 else 'left',
            )
        ws.row_dimensions[ri].height = 30

    col_widths = [5, 13, 28, 35, 12, 12, 22, 16, 16, 30, 16, 32]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'B4'

    # ── Лист 2: справочник ──
    wl = wb.create_sheet(title='Справочник')
    wl.merge_cells('A1:C1')
    wl['A1'].value     = 'Легенда цветов (итоги работы по обращению)'
    wl['A1'].font      = Font(bold=True, size=12, color=HEADER_COLOR)
    wl['A1'].alignment = Alignment(horizontal='center')
    wl.row_dimensions[1].height = 22

    for ci, h in enumerate(['Цвет', 'Итог', 'Обозначение'], 1):
        c = wl.cell(row=2, column=ci, value=h)
        c.fill = PatternFill('solid', fgColor=HEADER_COLOR)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.border = _std_border()
        c.alignment = Alignment(horizontal='center', vertical='center')
    wl.row_dimensions[2].height = 20

    if result_types:
        for li, rt in enumerate(result_types, 3):
            argb = _hex_to_argb(rt['color_hex'] or '')
            fill = PatternFill('solid', fgColor=argb) if argb else PatternFill('solid', fgColor='FFFFFFFF')
            ca = wl.cell(row=li, column=1, value='')
            ca.fill = fill; ca.border = _std_border()
            cb = wl.cell(row=li, column=2, value=rt['name'])
            cb.fill = fill
            cb.font = Font(bold=True, size=10); cb.border = _std_border()
            cb.alignment = Alignment(vertical='center')
            cc = wl.cell(row=li, column=3, value=rt['color_hex'])
            cc.font = Font(italic=True, size=9, color='888888'); cc.border = _std_border()
            cc.alignment = Alignment(vertical='center')
            wl.row_dimensions[li].height = 18
    else:
        wl.cell(row=3, column=1,
                value='Справочник итогов пуст. Добавьте значения в разделе «Справочники».')

    wl.column_dimensions['A'].width = 8
    wl.column_dimensions['B'].width = 36
    wl.column_dimensions['C'].width = 12

    fn = f"minek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fp = os.path.join(REPORTS_DIR, fn)
    wb.save(fp)

    # ── Журнал ──
    detail = f"период: {_fmt_date(df)} – {_fmt_date(dt)}; всего {len(rows)} обращ."
    try:
        log_action(conn, session['user_id'], 'export_minek', detail=detail)
        conn.commit()
    except Exception:
        pass
    conn.close()

    return send_file(fp, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── ПОЛНАЯ ВЫГРУЗКА БАЗЫ (для дозаполнения и импорта) ────────────────────

@report_bp.route('/export/full')
@login_required
def export_full():
    conn = get_db()
    rows = conn.execute("""
        SELECT r.*,
               ass.full_name AS assigned_name,
               st.name       AS subject_type_name,
               rt.name       AS result_type_name
        FROM requests r
        LEFT JOIN users         ass ON r.assigned_to     = ass.id
        LEFT JOIN subject_types st  ON r.subject_type_id = st.id
        LEFT JOIN result_types  rt  ON r.result_type_id  = rt.id
        ORDER BY r.id
    """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'База обращений'

    hfill = PatternFill("solid", fgColor="1B5E7B")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    id_fill = PatternFill("solid", fgColor="2E4057")  # тёмный фон для ID-колонки
    br    = _std_border()

    COLS = [
        ('id',                   'ID (не менять)'),
        ('request_number',       '№ обращения'),
        ('request_date',         'Дата обращения'),
        ('status',               'Статус'),
        ('applicant_full_name',  'Полное наименование'),
        ('applicant_short_name', 'Краткое наименование'),
        ('applicant_inn',        'ИНН'),
        ('project_name',         'Название проекта'),
        ('contact_person',       'Контактное лицо'),
        ('contact_phone',        'Телефон'),
        ('contact_email',        'E-mail'),
        ('investment_total',     'Инвестиции (млн руб.)'),
        ('jobs_total',           'Рабочих мест'),
        ('site_area_ha',         'Площадь (га)'),
        ('site_build_area_m2',   'Застройка (м²)'),
        ('preferred_districts',  'Районы'),
        ('source_type',          'Источник'),
        ('assigned_name',        'Ответственный'),
        ('subject_type_name',    'Предмет обращения'),
        ('feedback_date',        'Дата обратной связи'),
        ('result_type_name',     'Итоги работы'),
        ('incoming_number',      'Входящий номер'),
        ('answer_date',          'Дата ответа'),
        ('answer_method',        'Способ ответа'),
        ('answer_notes',         'Примечания к ответу'),
        ('additional_info',      'Доп. информация'),
    ]

    for ci, (field, header) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.fill = id_fill if field == 'id' else hfill
        c.font = hfont
        c.border = br
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = 'A2'

    keys = [r[0] for r in conn.execute('PRAGMA table_info(requests)').fetchall()]
    # добавляем join-поля
    extra_keys = ['assigned_name', 'subject_type_name', 'result_type_name']

    for ri, r in enumerate(rows, 2):
        row_keys = list(r.keys())
        for ci, (field, _) in enumerate(COLS, 1):
            val = r[field] if field in row_keys else None
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = br
            c.alignment = Alignment(vertical='center', wrap_text=(ci == len(COLS)))

    col_widths = [8, 16, 14, 12, 35, 25, 14, 30, 22, 16, 24, 14, 12, 10, 12, 24, 16, 20, 22, 14, 28, 18, 14, 18, 28, 30]
    for ci, w in enumerate(col_widths[:len(COLS)], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    fn = f"sonar_full_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fp = os.path.join(REPORTS_DIR, fn)
    wb.save(fp)

    log_action(conn, session['user_id'], 'export_full',
               detail=f'Полная выгрузка базы: {len(rows)} обращений')
    conn.commit()
    conn.close()

    return send_file(fp, as_attachment=True, download_name=fn,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── ИМПОРТ ОБНОВЛЁННОГО EXCEL ─────────────────────────────────────────────

@report_bp.route('/import/full', methods=['POST'])
@login_required
def import_full():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Только для администратора'}), 403

    file = request.files.get('import_file')
    if not file or not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Загрузите файл .xlsx'}), 400

    overwrite = request.form.get('overwrite') == '1'

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'Ошибка чтения файла: {e}'}), 400

    headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    COL_MAP = {
        'Дата обращения':        'request_date',
        'Полное наименование':   'applicant_full_name',
        'Краткое наименование':  'applicant_short_name',
        'ИНН':                   'applicant_inn',
        'Название проекта':      'project_name',
        'Контактное лицо':       'contact_person',
        'Телефон':               'contact_phone',
        'E-mail':                'contact_email',
        'Инвестиции (млн руб.)': 'investment_total',
        'Рабочих мест':          'jobs_total',
        'Площадь (га)':          'site_area_ha',
        'Застройка (м²)':        'site_build_area_m2',
        'Районы':                'preferred_districts',
        'Источник':              'source_type',
        'Дата обратной связи':   'feedback_date',
        'Входящий номер':        'incoming_number',
        'Дата ответа':           'answer_date',
        'Способ ответа':         'answer_method',
        'Примечания к ответу':   'answer_notes',
        'Доп. информация':       'additional_info',
    }
    FK_MAP = {
        'Предмет обращения': ('subject_type_id', 'subject_types'),
        'Итоги работы':      ('result_type_id',  'result_types'),
        'Ответственный':     ('assigned_to',      'users'),
    }

    try:
        id_idx = headers.index('ID (не менять)')
    except ValueError:
        return jsonify({'error': 'Колонка «ID (не менять)» не найдена. Используйте файл из «Скачать базу (Excel)»'}), 400

    conn = get_db()

    subjects  = {r['name']: r['id'] for r in conn.execute('SELECT id,name FROM subject_types').fetchall()}
    results   = {r['name']: r['id'] for r in conn.execute('SELECT id,name FROM result_types').fetchall()}
    users_map = {r['full_name']: r['id'] for r in conn.execute('SELECT id,full_name FROM users').fetchall()}
    fk_lookup = {
        'subject_type_id': subjects,
        'result_type_id':  results,
        'assigned_to':     users_map,
    }

    updated = 0
    skipped = 0
    errors  = []
    now     = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_id = row[id_idx]
        if not raw_id:
            continue
        try:
            rid = int(raw_id)
        except (ValueError, TypeError):
            errors.append(f'Невалидный ID: {raw_id}')
            continue

        existing = conn.execute('SELECT * FROM requests WHERE id=?', (rid,)).fetchone()
        if not existing:
            errors.append(f'ID {rid}: обращение не найдено в базе')
            continue

        updates = {}

        for ci, header in enumerate(headers):
            if ci == id_idx:
                continue
            cell_val = row[ci]

            if header in COL_MAP:
                field = COL_MAP[header]
                if cell_val is None or str(cell_val).strip() == '':
                    continue
                val = str(cell_val).strip()
                if not overwrite and existing[field] not in (None, ''):
                    continue
                updates[field] = val

            elif header in FK_MAP:
                field, _ = FK_MAP[header]
                if cell_val is None or str(cell_val).strip() == '':
                    continue
                name = str(cell_val).strip()
                lookup = fk_lookup[field]
                fk_id = lookup.get(name)
                if fk_id is None:
                    errors.append(f'ID {rid}: «{name}» не найдено в справочнике «{header}»')
                    continue
                if not overwrite and existing[field] not in (None, ''):
                    continue
                updates[field] = fk_id

        if not updates:
            skipped += 1
            continue

        set_clause = ', '.join(f'{k}=?' for k in updates)
        vals = list(updates.values()) + [now, session['user_id'], rid]
        conn.execute(
            f'UPDATE requests SET {set_clause}, updated_at=?, updated_by=? WHERE id=?',
            vals
        )
        log_action(conn, session['user_id'], 'import_xlsx', rid,
                   f'Импорт Excel: обновлены поля: {", ".join(updates.keys())}')
        updated += 1

    conn.commit()
    conn.close()

    return jsonify({'updated': updated, 'skipped': skipped, 'errors': errors})


# ─── AUTOSAVE / WAL CHECKPOINT ───────────────────────────────────────────────

@report_bp.route('/autosave', methods=['POST'])
@login_required
def autosave():
    try:
        conn = get_db()
        conn.execute('PRAGMA wal_checkpoint(PASSIVE)')
        conn.close()
        return jsonify({'status': 'ok'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500
